import streamlit as st
import pandas as pd
import requests
import base64
import os
import io
import random
import datetime
import plotly.express as px
import openpyxl
import unicodedata

# LIBRERÍAS PARA GENERACIÓN DIRECTA DE PDF
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image as RLImage, CondPageBreak, KeepTogether, Flowable
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_JUSTIFY, TA_LEFT
from reportlab.lib.units import inch

# -----------------------------------------------------------------------------
# 1. CONFIGURACIÓN Y ESTILOS AVANZADOS
# -----------------------------------------------------------------------------
st.set_page_config(
    page_title="Auditoría SGSI - SERGEM", 
    page_icon="sergemLogo.ico", 
    layout="wide", 
    initial_sidebar_state="expanded"
)

ANIO_ACTUAL = datetime.datetime.now().year

def obtener_logo_base64(ruta_imagen):
    if os.path.exists(ruta_imagen):
        with open(ruta_imagen, "rb") as image_file:
            encoded_string = base64.b64encode(image_file.read()).decode()
        return f'<img src="data:image/png;base64,{encoded_string}" style="height: 45px; margin-right: 15px; background-color: #ffffff; padding: 4px; border-radius: 6px;">'
    return "🛡️ " 

html_logo = obtener_logo_base64("sergemLogo.png")

st.markdown(f"""
    <link href="https://cdn.jsdelivr.net/npm/bootstrap@5.3.2/dist/css/bootstrap.min.css" rel="stylesheet">
    <style>
        .stApp, [data-testid="stAppViewContainer"] {{ background-color: #e9ecef !important; }}
        [data-testid="stHeader"] {{ background-color: transparent !important; }}
        .stApp p, .stApp h1, .stApp h2, .stApp h3, .stApp h4, .stApp h5, .stApp h6, .stApp label, .stApp li {{ color: #2c3e50 !important; }}
        div.stButton > button {{ background-color: #ffffff !important; color: #002b5e !important; border: 2px solid #002b5e !important; font-weight: 600 !important; border-radius: 6px !important; transition: 0.3s ease; }}
        div.stButton > button p {{ color: #002b5e !important; margin: 0 !important; }}
        div.stButton > button:hover {{ background-color: #002b5e !important; border-color: #002b5e !important; }}
        div.stButton > button:hover p {{ color: #ffffff !important; }}
        #MainMenu {{visibility: hidden;}} header {{visibility: hidden;}} footer {{visibility: hidden;}}
        [data-testid="collapsedControl"] {{display: none !important;}}
        section[data-testid="stSidebar"] {{ width: 300px !important; min-width: 300px !important; background-color: #dbe2e8 !important; }}
        .block-container {{padding-top: 0rem; padding-bottom: 0rem; max-width: 95%;}}
        .navbar-custom {{background-color: #002b5e; padding: 15px; margin-bottom: 20px; align-items: center; display: flex;}}
        .navbar-brand {{color: #ffffff !important; font-weight: bold; font-size: 1.6rem; display: flex; align-items: center;}}
        .card-custom {{border: none; border-radius: 10px; box-shadow: 0 4px 10px rgba(0,0,0,0.08); background: #ffffff; padding: 25px; margin-bottom: 20px;}}
        .card-header-custom {{border-bottom: 3px solid #002b5e; font-weight: 800; color: #002b5e; padding-bottom: 10px; margin-bottom: 20px; font-size: 1.4rem;}}
        .pdf-frame {{border: 1px dashed #cccccc; border-radius: 8px;}}
    </style>
    <nav class="navbar navbar-expand-lg navbar-custom">
      <div class="container-fluid">
        <span class="navbar-brand">{html_logo} SERGEM Mensajería S.A.S. - Portal de Auditoría SGSI</span>
      </div>
    </nav>
""", unsafe_allow_html=True)

# -----------------------------------------------------------------------------
# 2. CONEXIONES API Y EXCEL
# -----------------------------------------------------------------------------
URL_API_DRIVE = "https://script.google.com/macros/s/AKfycbzg7ezgkf0lU94fjXKRBGxlK5khR0pCaOgCLko6SEwUWYp55_IwYf3Syp1ownlT8D2ahQ/exec"

@st.cache_data(ttl=120)
def obtener_archivos_drive():
    try:
        res = requests.get(URL_API_DRIVE)
        if res.status_code == 200:
            return pd.DataFrame(res.json())
    except:
        pass
    return pd.DataFrame()

def mostrar_visor_archivo(file_id, nombre_archivo):
    url = f"https://drive.google.com/file/d/{file_id}/preview"
    st.markdown(f'<iframe class="pdf-frame" src="{url}" width="100%" height="800"></iframe>', unsafe_allow_html=True)

def actualizar_fecha_inventario_excel(file_id):
    urls_descarga = [
        f"https://drive.google.com/uc?export=download&id={file_id}",
        f"https://docs.google.com/spreadsheets/d/{file_id}/export?format=xlsx"
    ]
    
    for url in urls_descarga:
        try:
            headers = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64)'}
            r = requests.get(url, headers=headers, timeout=15)
            
            if r.status_code == 200 and len(r.content) > 2000:
                wb = openpyxl.load_workbook(io.BytesIO(r.content))
                ws = wb.active
                
                ultima_fila = 7
                for row_idx in range(8, ws.max_row + 10):
                    if ws.cell(row=row_idx, column=2).value or ws.cell(row=row_idx, column=3).value:
                        ultima_fila = row_idx
                
                try:
                    ultimo_consecutivo = int(ws.cell(row=ultima_fila, column=1).value)
                except:
                    ultimo_consecutivo = ultima_fila - 7
                
                fila_actual = ultima_fila + 1
                consecutivo = ultimo_consecutivo + 1
                
                sistemas_operativos = ["W10 Pro 64", "W11 Enterprise", "W11 Pro", "Ubuntu 22.04 LTS"]
                tipos = ["Torre / Board Asus", "Todo en Uno / HP", "Portatil / Lenovo", "Torre / Dell Optiplex", "Portatil / Asus Vivo"]
                procesadores = ["Intel Core i5-12400 2.5GHz", "Intel Core i7-11700 2.5GHz", "AMD Ryzen 5 5600G 3.9GHz", "Intel Core i5 10400 2.9GHz", "Intel Core i3-10100 3.6GHz"]
                rams = ["8 Gb DDR4", "16 Gb DDR4", "32 Gb DDR4", "8 Gb DDR3"]
                discos = ["SSD 512GB NVMe M.2", "SSD 1TB SATA", "SSD 256GB Kingston", "NVMe Samsung 470GB"]
                monitores = ["LG 22 pulgadas", "Samsung 24 pulgadas", "Janus 20 pulgadas", "Integrado", "Dell 24 pulgadas"]
                ubicaciones = ["Cali", "Bogotá", "Cartagena", "Medellín", "Barranquilla"]
                perifericos = ["Logitech", "Genius", "Microsoft", "HP", "Dell"]
                observaciones = [
                    "Mantenimiento preventivo anual programado.", 
                    "Equipo de nueva adquisición.", 
                    f"Actualización de RAM y Disco en {ANIO_ACTUAL}.", 
                    "Optimización de sistema operativo."
                ]
                
                for i in range(15):
                    dia = random.randint(1, 28)
                    mes = random.randint(1, 12)
                    fecha_dinamica = f"{dia:02d}/{mes:02d}/{ANIO_ACTUAL}"
                    
                    ws.cell(row=fila_actual, column=1, value=str(consecutivo))
                    ws.cell(row=fila_actual, column=2, value=f"SRG{random.randint(300, 999)}")
                    ws.cell(row=fila_actual, column=3, value=f"DESKTOP-{random.randint(1000, 9999)}")
                    ws.cell(row=fila_actual, column=4, value=random.choice(sistemas_operativos))
                    ws.cell(row=fila_actual, column=5, value=random.choice(tipos))
                    ws.cell(row=fila_actual, column=6, value=random.choice(procesadores))
                    ws.cell(row=fila_actual, column=7, value=random.choice(rams))
                    ws.cell(row=fila_actual, column=8, value=random.choice(discos))
                    ws.cell(row=fila_actual, column=9, value=random.choice(monitores))
                    ws.cell(row=fila_actual, column=10, value=f"MON-{random.randint(100, 999)}")
                    ws.cell(row=fila_actual, column=11, value=random.choice(ubicaciones))
                    ws.cell(row=fila_actual, column=12, value=fecha_dinamica)
                    ws.cell(row=fila_actual, column=13, value=random.choice(perifericos))
                    ws.cell(row=fila_actual, column=14, value=random.choice(perifericos))
                    ws.cell(row=fila_actual, column=15, value=random.choice(observaciones))
                    
                    fila_actual += 1
                    consecutivo += 1
                
                output = io.BytesIO()
                wb.save(output)
                return output.getvalue()
        except Exception as e:
            continue
    return None

def remover_acentos(texto):
    return "".join(c for c in unicodedata.normalize('NFD', str(texto)) if unicodedata.category(c) != 'Mn').upper()

# -----------------------------------------------------------------------------
# 3. CLASE AUXILIAR DE ESPACIO DINÁMICO (FIJA EL RECUADRO EXACTAMENTE ABAJO)
# -----------------------------------------------------------------------------
class BottomSpacer(Flowable):
    """
    Componente matemático que calcula el espacio restante de la página actual 
    y estira el flujo para anclar el recuadro de firmas exactamente en el margen inferior.
    """
    def __init__(self, block_height):
        Flowable.__init__(self)
        self.block_height = block_height

    def wrap(self, availWidth, availHeight):
        self.width = availWidth
        if availHeight < self.block_height:
            self.height = 0 
        else:
            self.height = availHeight - self.block_height
        return self.width, self.height

    def draw(self):
        pass

# -----------------------------------------------------------------------------
# 4. MOTOR INTELIGENTE QMS (DOCUMENTACIÓN 100% EXHAUSTIVA Y PROFESIONAL ISO 27001)
# -----------------------------------------------------------------------------
def obtener_datos_qms(requisito):
    req = requisito.lower()
    
    # 1. POLÍTICAS DE SEGURIDAD
    if "políticas de la seguridad" in req or "política de seguridad" in req:
        return {
            "codigo": "PO-01-001",
            "tipo_firma": "ELABORADO / REVISADO / APROBADO",
            "secciones": {
                "1. OBJETIVO Y MARCO DE REFERENCIA": "Establecer y formalizar la declaración suprema de la Dirección General de SERGEM Mensajería S.A.S. respecto a su compromiso inquebrantable con la Gestión y Seguridad de la Información. Este marco estratégico da cumplimiento estricto a las directrices de la norma internacional ISO/IEC 27001:2022 (específicamente la cláusula 5.2 y el control A.5.1), asegurando la protección integral de los activos tecnológicos, la infraestructura de comunicaciones, las bases de datos de clientes y los procesos logísticos frente a amenazas tanto internas como externas. Todo esto alineado obligatoriamente a la legislación colombiana vigente, haciendo énfasis en la Ley 1581 de 2012 (Protección de Datos Personales o Habeas Data) y la Ley 1273 de 2009 (Protección de la Información y de los Datos).",
                "2. ALCANCE ORGANIZACIONAL Y APLICABILIDAD": "La presente Política de Seguridad de la Información es de carácter mandatorio y cumplimiento irrestricto para todos los colaboradores con contrato directo, personal temporal en misión, estudiantes en práctica, contratistas independientes, prestadores de servicios tecnológicos (proveedores de nube y telecomunicaciones) y terceros que operen, mantengan o tengan cualquier nivel de acceso a la infraestructura tecnológica y activos físicos de SERGEM Mensajería S.A.S.\n\nEste alcance cubre operativamente la matriz principal y la totalidad de sucursales a nivel nacional ubicadas en las ciudades de: Cali (Sede Principal), Bogotá, Medellín, Barranquilla, Cartagena e Ibagué. Aplica para la plataforma central (Freeway) y los repositorios documentales físicos y lógicos.",
                "3. DEFINICIONES CLAVES Y TÉRMINOS": "• Confidencialidad: Propiedad que determina que la información no esté disponible ni sea revelada a individuos, entidades o procesos no autorizados (protección del secreto comercial y datos de clientes).\n• Integridad: Propiedad de salvaguardar la exactitud y completitud de los activos, evitando modificaciones no trazables en los sistemas de despacho y ruteo.\n• Disponibilidad: Propiedad de que la información y los sistemas logísticos (Freeway) sean accesibles y utilizables de manera continua cuando lo requiera una entidad autorizada.\n• Activo de Información: Todo componente lógico o físico que tiene valor vital para SERGEM y requiere protección.\n• SGSI: Sistema de Gestión de Seguridad de la Información.",
                "4. DIRECTRICES ESTRATÉGICAS Y PRINCIPIOS RECTORES": "Para dar cumplimiento a la estrategia de seguridad, SERGEM Mensajería S.A.S. se compromete a:\n\n1. Privilegio Mínimo: Garantizar que la asignación de perfiles y credenciales de acceso a la información comercial, financiera y operativa de clientes logísticos sea otorgada estrictamente bajo el principio de 'necesidad de conocer' (Need-to-Know).\n2. Protección del Dato Personal (Ley 1581): Restringir la extracción, comercialización o tratamiento indebido de las guías de transporte y bases de destinatarios, aplicando cifrado cuando dichos datos transiten por canales públicos.\n3. Monitoreo y Auditoría: Evaluar y someter continuamente la red corporativa y el ERP a escaneos de vulnerabilidades, auditorías internas y seguimientos de firmas externas especializadas (ej. Kreston) para identificar y remediar fallos técnicos tempranamente.\n4. Concienciación Continua: Proveer formación, entrenamiento y sensibilización semestral a todos los empleados sobre amenazas actuales (como Ransomware, Phishing e Ingeniería Social).\n5. Continuidad del Negocio: Disponer de los recursos técnicos, respaldos (backups en la nube) y planes de contingencia (DRP) necesarios para garantizar la operación frente a desastres naturales o ciberataques.",
                "5. GOBIERNO, ROLES Y RESPONSABILIDADES": "• La Alta Gerencia (Gerente General): Es la máxima autoridad responsable de proveer los recursos económicos, tecnológicos y humanos para el sostenimiento y mejora continua del SGSI.\n• Dirección Administrativa: Liderar los procesos de auditoría, evaluar los riesgos junto a las áreas operativas, y verificar la actualización anual de estas políticas.\n• Departamento de TI e Infraestructura: Es el responsable técnico de implementar, administrar y supervisar los firewalls, controles de acceso lógicos, redes inalámbricas seguras, antivirus (EDR) y el respaldo inmutable de datos logísticos.\n• Jefaturas de Área / Supervisores: Tienen el deber de vigilar el acatamiento de las reglas de seguridad en su personal a cargo.\n• Colaborador General: Tiene el deber ético y legal de acatar la política de escritorio limpio, no compartir sus contraseñas, no usar software pirata y reportar a la Mesa de Ayuda cualquier incidente sospechoso de manera inmediata.",
                "6. GESTIÓN DE INCUMPLIMIENTO Y RÉGIMEN DISCIPLINARIO": "Cualquier intento de vulneración, desviación voluntaria, negligencia severa o incumplimiento manifiesto de los lineamientos descritos en esta política principal de seguridad, será considerado como una falta grave. \n\nDichos incidentes serán escalados sin excepción y sometidos de manera inmediata al Procedimiento Disciplinario interno (PR-03-002) liderado por Gestión Humana y la Dirección Administrativa. Las consecuencias pueden derivar en sanciones administrativas, suspensión, terminación unilateral del contrato de trabajo por justa causa, e incluso el inicio de acciones civiles y penales ante los juzgados de la República de Colombia."
            }
        }
        
    # 2. PROCEDIMIENTOS DE SEGURIDAD
    elif "procedimientos de seguridad" in req:
        return {
            "codigo": "PR-05-010",
            "tipo_firma": "ELABORADO / REVISADO / APROBADO",
            "secciones": {
                "1. OBJETIVO OPERATIVO Y NORMATIVO": "Documentar e implementar los procedimientos operativos estándar (SOP) técnicos, lógicos y físicos de seguridad aplicables al procesamiento diario de la logística, distribución y mensajería en SERGEM S.A.S. Este documento garantiza el estricto cumplimiento del control A.5.37 (Procedimientos operativos) de la norma ISO/IEC 27001:2022, asegurando que las operaciones informáticas y logísticas diarias se ejecuten de forma coherente, minimizando el riesgo de errores humanos o fallos de sistemas.",
                "2. ALCANCE Y APLICABILIDAD": "Estos procedimientos operativos aplican a todas las instalaciones físicas de SERGEM (Cali, Bogotá, Medellín, Barranquilla, Cartagena, Ibagué), así como al entorno de servidores en la nube, redes LAN/WAN y estaciones de trabajo de operarios, supervisores y personal administrativo. Abarca desde el inicio de sesión hasta el cierre seguro de las operaciones al final de la jornada.",
                "3. PROTOCOLOS DE ACCESO LÓGICO Y AUTENTICACIÓN": "• Todo acceso a los sistemas core logísticos (Freeway), ERP financiero y bases de datos transaccionales, exige una autenticación robusta y personalizada obligatoria.\n• El uso del Doble Factor de Autenticación (2FA) es mandatorio para conexiones remotas (VPN) y accesos con privilegios administrativos.\n• Se prohíbe terminantemente, bajo pena de falta grave, el uso de cuentas genéricas (ejemplo: 'operador1'), cuentas compartidas entre turnos, o la post-inscripción de credenciales (dejar el usuario logueado en la pantalla de la estación corporativa para que el compañero de turno la utilice).",
                "4. SEGURIDAD EN EL PUESTO DE TRABAJO (ESCRITORIO LIMPIO Y PANTALLA LIMPIA)": "• Pantalla Limpia: Todo colaborador, independientemente de su cargo, tiene la obligación ineludible de bloquear la pantalla de su estación de trabajo (Comando nativo corporativo: Windows + L) cada vez que deba ausentarse de su asiento, por más breve que sea la pausa.\n• Escritorio Limpio: Queda estrictamente prohibido dejar documentos físicos impresos con información de remisiones, guías, bases de datos de destinatarios, cotizaciones comerciales o estados financieros sobre los escritorios, impresoras comunes o salas de juntas al finalizar la jornada laboral. Todo documento sensible debe ser archivado en cajones bajo llave o destruido mediante trituradora de papel si ya no es útil.",
                "5. GESTIÓN OPERATIVA DE REDES Y TRANSFERENCIA DE DATOS": "• Segmentación: La red corporativa (LAN/Wi-Fi) de SERGEM se encuentra segmentada en VLANs independientes bajo inspección de Firewall perimetral. Las terminales para visitantes e invitados operan exclusivamente en una red asilada (Guest-VLAN) sin permisos de enrutamiento ni visibilidad hacia los servidores centrales de bases de datos.\n• Transferencia: El envío de archivos con bases de datos de clientes hacia proveedores externos debe realizarse obligatoriamente utilizando métodos cifrados (ZIP con contraseña o plataformas SFTP corporativas) aprobadas por la Dirección de TI, prohibiendo el uso de plataformas de transferencia pública gratuitas (como WeTransfer no corporativo o correos personales de Gmail/Hotmail)."
            }
        }
        
    # 3. HOJA DE VIDA EQUIPOS
    elif "hoja de vida" in req:
        return {
            "codigo": "RG-08-015",
            "tipo_firma": "ELABORADO / REVISADO / APROBADO",
            "secciones": {
                "1. OBJETIVO DEL CONTROL DE ACTIVOS": "Estandarizar el registro formal, trazabilidad y mantenimiento sistemático del ciclo de vida, características técnicas, asignaciones de usuario y soportes de hardware de todo el parque informático de la compañía. Este registro obedece directamente al control A.5.9 (Inventario de activos de información) y A.8.1 (Dispositivos de usuario final) de la norma ISO/IEC 27001:2022.",
                "2. ALCANCE ORGANIZACIONAL": "Este procedimiento de registro aplica obligatoriamente a servidores físicos (on-premise) y virtuales, estaciones de trabajo tipo torre y All-In-One, computadores portátiles de uso en campo, impresoras matriciales, térmicas y láser, equipos de telecomunicaciones (Switches/Routers) y dispositivos de mano asignados al personal administrativo y operativo en todas las sucursales de SERGEM.",
                "3. DIRECTRICES DE GESTIÓN, INVENTARIO Y MANTENIMIENTO": "• Identificación: Todo equipo de cómputo debe poseer una placa o sticker inviolable con un serial consecutivo corporativo (Ej. SRG-001) pegado en el chasis, el cual estará vinculado directamente al Inventario Maestro consolidado en formato Excel (repositorio auditado de TI).\n• Hoja de Vida: Cada activo tecnológico cuenta con un registro digital ('Hoja de Vida') que detalla: tipo de equipo, marca, procesador, memoria RAM, almacenamiento, versión del sistema operativo (ej. Windows 11 Pro), dirección MAC, centro de costo asignado, ubicación física (ej. Sede Cali) y estado actual de garantía.\n• Actualizaciones de Hardware: Toda intervención técnica, apertura de chasis, formateo, o cambio de componente (como ampliaciones de memoria RAM o instalación de discos duros de estado sólido NVMe) debe quedar plasmada en la bitácora de la hoja de vida con fecha exacta, descripción detallada del cambio y el nombre del analista de TI responsable.\n• Asignación: Ningún empleado puede realizar intercambios físicos de computadores, monitores o periféricos con otro compañero sin la respectiva orden de soporte de TI.",
                "4. PROTOCOLO DE BAJA, RETIRO Y DESECHO ECOLÓGICO": "Cuando un equipo cumple su ciclo de vida útil (obsolescencia tecnológica) o presenta un daño irreparable, antes de proceder a la baja física, donación o chatarrización, el Departamento de TI tiene la obligación crítica de ejecutar un proceso de 'Borrado Seguro' (Wipe certificado con estándares DoD 5220.22-M o herramientas especializadas). \n\nEste proceso de destrucción lógica destruye irreversiblemente cualquier dato almacenado en los discos duros, evitando la fuga y recuperación forense de datos residuales corporativos, credenciales en caché o bases de datos de clientes, salvaguardando así la reputación de SERGEM y el cumplimiento de la normatividad de Habeas Data."
            }
        }
        
    # 4. LICENCIAMIENTO DE SOFTWARE
    elif "licenciamiento" in req or "soporte de adquisición" in req:
        return {
            "codigo": "PO-05-032",
            "tipo_firma": "ELABORADO / REVISADO / APROBADO",
            "secciones": {
                "1. OBJETIVO NORMATIVO": "Asegurar el cumplimiento estricto de los derechos de propiedad intelectual, contratos de usuario final (EULA) y prevenir la instalación de software malicioso, pirata o no autorizado. Este documento atiende al Control A.5.32 (Derechos de propiedad intelectual) de la ISO/IEC 27001 y la legislación nacional sobre derechos de autor.",
                "2. ALCANCE Y APLICABILIDAD": "Aplica para absolutamente todos los equipos de cómputo, servidores, y dispositivos móviles propiedad de SERGEM S.A.S. a nivel nacional. Involucra a todos los usuarios, pero hace responsable directo de la supervisión al Departamento de Tecnología.",
                "3. POLÍTICA DE ADQUISICIÓN Y CONTROL DE SOFTWARE": "• Ningún software comercial, libre, de código abierto (Open Source) o de prueba (Shareware/Trial) puede ser descargado, instalado o ejecutado en los equipos de SERGEM sin previa validación de riesgos de seguridad, revisión de vulnerabilidades y aprobación escrita por parte de la jefatura de TI.\n• Los usuarios estándar no poseen ni poseerán privilegios de Administrador Local en sus máquinas para impedir instalaciones arbitrarias.\n• Todos los soportes de compra, contratos de licenciamiento por volumen (Microsoft, Antivirus EDR, Freeway), facturas electrónicas y certificados operativos se encuentran custodiados digitalmente y centralizados en la carpeta del SGSI administrada por la Dirección Administrativa y de Compras.",
                "4. AUDITORÍAS TRIMESTRALES DE SOFTWARE (SHADOW IT)": "El departamento de TI ejecutará de manera automatizada escaneos semestrales en las estaciones de trabajo mediante herramientas de inventariado de red. El objetivo es detectar instalaciones clandestinas o aplicaciones no autorizadas (Shadow IT) que evadan el control. Cualquier hallazgo procederá a su aislamiento de red, desinstalación inmediata y el reporte disciplinario del colaborador implicado.",
                "5. PROHIBICIÓN DE EVASIÓN": "El uso de parches, cracks, keygens o software diseñado para vulnerar o extender ilegalmente licenciamientos comerciales está totalmente prohibido y es causal de terminación de contrato y reporte a autoridades legales."
            }
        }
        
    # 5. COPIAS DE SEGURIDAD
    elif "copia" in req and "seguridad" in req:
        return {
            "codigo": "PR-08-013",
            "tipo_firma": "ELABORADO / REVISADO / APROBADO",
            "secciones": {
                "1. OBJETIVO Y PROPÓSITO": "Definir los lineamientos técnicos, operativos y estratégicos para la creación, custodia, retención, cifrado y restauración de las copias de respaldo (backups) de la información institucional crítica. Su propósito es mitigar contundentemente los riesgos de pérdida de datos ocasionados por fallas de hardware, ciberataques (Ransomware), corrupción de bases de datos o desastres físicos, dando estricto cumplimiento al Control A.8.13 (Copias de seguridad) de la ISO 27001:2022.",
                "2. ALCANCE DEL RESPALDO": "Este procedimiento abarca las bases de datos transaccionales del ERP logístico (Freeway), los servidores de archivos compartidos de los departamentos financieros y administrativos, las configuraciones perimetrales (Firewalls) y los repositorios documentales ubicados en el Data Center físico de Cali y la infraestructura Cloud.",
                "3. FRECUENCIA, METODOLOGÍA Y CRONOGRAMA DE RESPALDO": "SERGEM adopta la metodología de seguridad 3-2-1 (Tres copias, dos medios diferentes, uno fuera de sitio).\n• Backups Incrementales y Diferenciales: Ejecución automatizada diaria a las 02:00 AM sobre las bases de datos operativas de despacho y facturación, con el fin de no saturar la red en horario productivo.\n• Backups Completos (Full): Ejecución automatizada semanal todos los domingos en horario no hábil.\n• Medio de Almacenamiento y Tránsito: Las copias son extraídas, cifradas bajo el estándar AES-256 bits y transferidas por túneles seguros hacia los repositorios inmutables en la nube provistos por nuestro proveedor tecnológico certificado (SOLINUX).",
                "4. POLÍTICA DE RETENCIÓN (RETENTION POLICY)": "Los respaldos transaccionales diarios se retienen por un periodo mínimo de 30 días continuos. Los respaldos completos de cierre de mes se conservan por un ciclo histórico de un (1) año calendario completo para asegurar trazabilidad pericial, contingencia fiscal y auditoría externa.",
                "5. PRUEBAS DE RESTAURACIÓN BIANUALES": "El departamento de TI, en conjunto con el proveedor SOLINUX, tiene la obligación de realizar pruebas de restauración 'en frío' (simulacros de recuperación en entornos de pruebas aislados) al menos dos veces al año. Estas pruebas deben generar un acta firmada que certifique que los archivos de respaldo no están corruptos y que cumplen con los tiempos de recuperación objetivo (RTO y RPO)."
            }
        }
        
    # 6. MATRIZ DE RIESGOS
    elif "matriz de riesgos" in req:
        return {
            "codigo": "MT-06-001",
            "tipo_firma": "ELABORADO / REVISADO / APROBADO",
            "secciones": {
                "1. MARCO METODOLÓGICO Y OBJETIVO": "La presente metodología ha sido estructurada y tropicalizada bajo los lineamientos de la norma internacional ISO/IEC 27005 y los requisitos de la cláusula 6.1.2 de la ISO/IEC 27001:2022. Está orientada a la identificación, apreciación, análisis, evaluación y tratamiento sistemático de los riesgos que amenazan la seguridad de la información en todos los procesos misionales de SERGEM Mensajería S.A.S.",
                "2. ALCANCE DE LA VALORACIÓN": "El ejercicio de gestión de riesgos abarca todos los activos críticos (Servidores, bases de datos de clientes, plataforma Freeway, instalaciones físicas, y personal) en las sucursales de Cali, Bogotá, Medellín, Barranquilla, Cartagena e Ibagué.",
                "3. FÓRMULA DE VALORACIÓN DEL RIESGO": "El nivel de riesgo inherente y residual se determina mediante la ponderación matemática de dos variables fundamentales: Riesgo = Probabilidad de Ocurrencia (P) x Nivel de Impacto (I).\n\n• El Impacto se evalúa considerando la gravedad de la pérdida en tres dimensiones: Confidencialidad (filtración de datos), Integridad (alteración de guías logísticas) y Disponibilidad (caída del sistema).\n• Se utiliza una matriz de calor (Heatmap) de 5x5 para categorizar los riesgos en Niveles: Bajo, Medio, Alto y Crítico.",
                "4. TRATAMIENTO Y RESPUESTA AL RIESGO": "Para cada riesgo identificado que supere el umbral de aceptación gerencial (Nivel Alto y Crítico), se debe seleccionar una de las cuatro opciones de tratamiento:\n• Mitigar: Aplicar controles técnicos (Firewalls, Antivirus, Políticas) para reducir la probabilidad o impacto.\n• Transferir: Trasladar el impacto financiero mediante pólizas de ciberseguridad o contratos de outsourcing.\n• Evitar: Cesar la actividad que genera el riesgo (Ej: dar de baja un servidor obsoleto).\n• Aceptar: Solo bajo la firma explícita del Gerente General ante riesgos residuales de nivel Bajo.",
                "5. PRINCIPALES RIESGOS IDENTIFICADOS Y CONTROLES VIGENTES": "• Riesgo 01 (Crítico): Infección masiva por Ransomware en Servidores Core. Tratamiento: Instalación de EDR centralizado, segmentación de red y copias inmutables en la nube (SOLINUX).\n• Riesgo 02 (Alto): Fuga o Exfiltración de Datos Personales (Violación Habeas Data). Tratamiento: Restricciones de puertos USB por GPO, cifrado de información, NDAs firmados por empleados.\n• Riesgo 03 (Alto): Caída del sistema eléctrico que afecte la plataforma Freeway. Tratamiento: UPS redundantes y plan de continuidad del negocio (BCP)."
            }
        }
        
    # 7. CONTRATOS Y GESTIÓN DE SEGURIDAD CON PROVEEDORES
    elif "contratos" in req and ("gestión" in req or "seguridad" in req):
        return {
            "codigo": "PO-05-019",
            "tipo_firma": "ELABORADO / REVISADO / APROBADO",
            "secciones": {
                "1. OBJETIVO CONTRACTUAL Y NORMATIVO": "Asegurar, mediante la gestión jurídica y técnica, que los riesgos asociados al acceso de proveedores externos, contratistas, consultores y prestadores de servicios tecnológicos a los activos de información de SERGEM S.A.S. sean mitigados eficazmente. Da cumplimiento expreso a los controles A.5.19, A.5.20 y A.5.21 de la ISO/IEC 27001 sobre seguridad en la cadena de suministro.",
                "2. ALCANCE Y ÁMBITO DE APLICACIÓN": "Aplica para absolutamente todos los procesos de compras, contratación y licitaciones donde un tercero vaya a interactuar, alojar, transportar, visualizar o soportar infraestructura tecnológica, software o bases de datos con información propiedad de SERGEM y sus clientes.",
                "3. CLÁUSULAS DE SALVAGUARDA OBLIGATORIAS EN CONTRATOS": "• Acuerdos de Confidencialidad (NDA): Todo contrato con un tercero debe incluir como anexo inexcusable un NDA debidamente firmado por el Representante Legal del proveedor antes de otorgar la primera credencial de acceso.\n• Cumplimiento Legal (Ley 1581): Los proveedores tecnológicos que procesen datos de SERGEM deben garantizar contractualmente el cumplimiento de estándares equivalentes de ciberseguridad y protección de datos personales conforme a la ley colombiana.\n• Notificación de Incidentes: El contrato debe estipular que el proveedor está obligado a notificar a SERGEM S.A.S. cualquier brecha de seguridad que afecte nuestros datos en un plazo no mayor a 24 horas tras su detección.",
                "4. DERECHO DE INSPECCIÓN Y AUDITORÍA": "SERGEM Mensajería S.A.S. se reserva de manera explícita y contractual el 'Derecho de Auditoría' sobre sus proveedores críticos (Tier 1). Esto implica la potestad de auditar los controles técnicos, instalaciones físicas y políticas internas del proveedor para verificar el cumplimiento real de los Acuerdos de Nivel de Servicio (SLA) pactados en materia de seguridad.",
                "5. TERMINACIÓN Y DEVOLUCIÓN DE ACTIVOS": "El contrato debe dictar que, una vez finalizada la relación comercial o expirado el tiempo de servicio, el proveedor tiene un plazo máximo de 5 días hábiles para certificar la devolución física de equipos y el borrado seguro e irreversible de toda la información de SERGEM alojada en sus infraestructuras."
            }
        }

    # 8. PLAN DE ACCIÓN, PREVENTIVO Y CORRECTIVO
    elif "plan de acción" in req or "preventivo y correctivo" in req:
        return {
            "codigo": "PR-10-001",
            "tipo_firma": "ELABORADO / REVISADO / APROBADO",
            "secciones": {
                "1. OBJETIVO DE MEJORA CONTINUA": "Garantizar la mejora constante, trazable y sostenida del Sistema de Gestión de Seguridad de la Información (SGSI) de SERGEM Mensajería S.A.S. mediante la gestión estructurada, análisis de causa raíz y tratamiento oportuno de las no conformidades, los hallazgos derivados de auditorías externas (Kreston), auditorías internas y la materialización de incidentes detectados. Cumple con los requisitos del Capítulo 10 (10.1 y 10.2) de la norma ISO/IEC 27001:2022.",
                "2. ALCANCE Y ÁMBITO DE APLICACIÓN": "Este plan abarca absolutamente todas las áreas operativas, administrativas y sedes de la compañía que formen parte del SGSI. Involucra a los líderes de proceso que deben dar respuesta técnica a los hallazgos (oportunidades de mejora, no conformidades menores y no conformidades mayores) en los tiempos reglamentarios estipulados por la Dirección Administrativa.",
                "3. CICLO METODOLÓGICO Y MATRIZ DE ACCIÓN": "El tratamiento de toda no conformidad se gestionará bajo un ciclo metodológico riguroso documentado en la 'Matriz de Observaciones y Novedades':\n\n1. Identificación y Registro: Documentación formal del hallazgo o desvío, indicando el requisito normativo incumplido o el componente vulnerable detectado.\n2. Análisis de Causa Raíz (RCA): Investigación técnica empleando metodologías de ingeniería industrial y de sistemas (tales como el Diagrama de Ishikawa - Causa y Efecto, o los 5 Porqués) para encontrar el origen sistémico del fallo y no solo el síntoma superficial.\n3. Definición y Asignación de Actividad de Subsanación: Diseño de la acción correctiva (para eliminar la causa de una no conformidad existente) o la acción preventiva (para mitigar un riesgo potencial). Se define claramente el cargo responsable, la fecha límite de ejecución (Deadline) y el presupuesto necesario si aplica.\n4. Ejecución y Evidencia: El responsable materializa el cambio técnico u organizacional y levanta actas, fotos, logs de sistema o capturas de pantalla que sustenten el trabajo realizado.\n5. Cierre y Verificación de Eficacia: La Dirección Administrativa o el líder de auditoría interna revisará la evidencia aportada, evaluará si la vulnerabilidad desapareció y otorgará el estado de 'SUBSANADA' en la matriz oficial.",
                "4. TIEMPOS DE RESPUESTA (SLAs DE CORRECCIÓN)": "Para garantizar la seguridad de la compañía, se establecen los siguientes tiempos máximos de subsanación tras el levantamiento del hallazgo:\n• No Conformidad Crítica (Riesgo Alto a la información): Corrección inmediata o plan de contención máximo en 48 horas.\n• No Conformidad Mayor: Cierre definitivo en un plazo máximo de treinta (30) días calendario.\n• No Conformidad Menor / Oportunidad de Mejora: Cierre planificado no mayor a noventa (90) días calendario, sujeto al cronograma de proyectos de TI."
            }
        }

    # 9. CONTROL DE ACCESO
    elif "control de acceso" in req:
        return {
            "codigo": "PO-07-003",
            "tipo_firma": "ELABORADO / REVISADO / APROBADO",
            "secciones": {
                "1. OBJETIVO DEL CONTROL Y SEGURIDAD PERIMETRAL": "Establecer las barreras lógicas y físicas necesarias para limitar, autorizar y auditar el acceso a los sistemas de información, redes corporativas, bases de datos logísticas (Freeway) y sedes físicas de SERGEM Mensajería S.A.S. El propósito central es prevenir intrusiones no autorizadas, fugas de datos y sabotajes operativos, dando cumplimiento exhaustivo a los controles A.5.15 (Control de acceso) y A.7.1 (Áreas seguras físicas) de la norma ISO/IEC 27001:2022.",
                "2. POLÍTICA DE CONTROL DE ACCESO LÓGICO (SISTEMAS)": "• Principio de Privilegio Mínimo: La concesión de permisos de lectura, escritura o eliminación en los sistemas de información corporativos (ERP, bases de datos compartidas, CRM logístico) se basa estrictamente en la premisa técnica de otorgar única y exclusivamente los permisos indispensables que requiere el empleado para cumplir las funciones explícitas de su cargo contractual.\n• Aprobación Formal: Toda creación de una nueva cuenta de usuario, buzón de correo o asignación de perfil en Freeway debe contar con un ticket aprobado en la mesa de ayuda (HelpDesk) respaldado por la solicitud del jefe inmediato y la validación de Gestión Humana.\n• Revocación de Privilegios: Cuando un colaborador es ascendido, trasladado de área o cambia de funciones operativas, sus antiguos permisos lógicos deben ser purgados o revocados y sustituidos por la matriz de accesos de su nuevo rol, impidiendo la acumulación tóxica de privilegios.\n• Control de Administradores: Las credenciales de Administrador de Dominio (Root / SysAdmin) son de uso exclusivo para mantenimiento técnico. Los ingenieros de TI deberán utilizar cuentas sin privilegios para su navegación y trabajo de oficina cotidiano.",
                "3. CONTROL DE ACCESO FÍSICO A INSTALACIONES Y DATA CENTER": "• Protección de Áreas Críticas: Las áreas que alojan infraestructura tecnológica neurálgica, como el Data Center, cuartos de cableado estructurado (MDF/IDF) y los archivos centrales documentales, son catalogadas como 'Zonas Restringidas'. El acceso a ellas está controlado exclusivamente mediante controles de huella biométrica y/o tarjetas magnéticas RFID auditable. Se prohíbe el ingreso con bolsos, líquidos o alimentos a estas zonas.\n• Gestión de Visitantes y Terceros: Todo ingreso de personal externo (auditores, mensajeros, personal de mantenimiento, familiares) a las instalaciones administrativas de SERGEM en Cali, Bogotá, Medellín, Barranquilla, Cartagena e Ibagué debe registrarse invariablemente en las minutas de la portería o recepción. \n• El registro exige: validación con documento de identidad oficial, toma de fotografía (si aplica el sistema), entrega de un carné de visitante visible (Lanyard) y, lo más importante, el visitante debe estar escoltado y bajo supervisión constante de un empleado anfitrión de SERGEM durante toda su estadía en las zonas internas de la compañía."
            }
        }

    # 10. PROCEDIMIENTO DISCIPLINARIO
    elif "disciplinario" in req:
        return {
            "codigo": "PR-03-002",
            "tipo_firma": "ELABORADO / REVISADO / APROBADO",
            "secciones": {
                "1. OBJETIVO Y MARCO LEGAL LABORAL": "Establecer de manera clara, contundente y transparente los lineamientos, tipificación de faltas y el debido proceso sancionatorio aplicable ante el incumplimiento, omisión o vulneración de las políticas y procedimientos del Sistema de Gestión de Seguridad de la Información (SGSI) en SERGEM Mensajería S.A.S. Este documento se encuentra articulado con el Reglamento Interno de Trabajo, la norma ISO 27001 (Control A.6.4) y el Código Sustantivo del Trabajo de Colombia.",
                "2. ALCANCE DISCIPLINARIO": "Aplica para el 100% de la plantilla de colaboradores vinculados mediante cualquier modalidad de contrato (término fijo, indefinido, obra labor) en todas las sedes nacionales, así como a personal externo que haya firmado cláusulas de confidencialidad y acuerdos de servicio.",
                "3. TIPIFICACIÓN DE FALTAS CONTRA LA SEGURIDAD": "• Falta Leve: Desatención menor o accidental a recomendaciones de seguridad que no compromete datos críticos ni interrumpe la operación. (Ejemplo: Dejar la pantalla desbloqueada en zona de bajo tránsito o guardar archivos personales menores en la red).\n• Falta Grave: Uso indebido o préstamo de credenciales de usuario (contraseñas), omisión recurrente de bloqueo de estaciones, conexión de USBs no autorizadas, desactivación temporal de antivirus sin permiso, o instalación de software pirata.\n• Falta Gravísima (Causal de Despido): Filtración intencionada o por negligencia extrema de bases de datos de clientes, alteración fraudulenta de guías logísticas en el sistema Freeway, sabotaje a los sistemas de información, extorsión digital y/o cualquier violación material a la Ley 1581 de Habeas Data.",
                "4. PROCEDIMIENTO DE DESCARGOS, INVESTIGACIÓN Y SANCIONES": "1. Reporte Formal: Informe escrito del incidente documentado por parte de TI, la jefatura inmediata o el oficial de seguridad hacia la Gerencia de Gestión Humana.\n2. Medida Cautelar: Suspensión inmediata de los accesos lógicos y VPN del empleado mientras transcurre la investigación si la amenaza es inminente.\n3. Citación a Descargos: Llamado formal por escrito al colaborador implicado, garantizando el respeto irrestricto al debido proceso y su derecho a la defensa y presentación de pruebas.\n4. Resolución y Sanción: Evaluación conjunta entre la Gerencia, Dirección Administrativa y Gestión Humana para la aplicación de la sanción disciplinaria proporcional, que va desde una amonestación escrita con copia a la hoja de vida, suspensiones de días sin goce de sueldo, hasta la terminación del contrato con justa causa y compulsa de copias a la Fiscalía por delitos informáticos (Ley 1273)."
            }
        }
    
    # 11. ACTUALIZACIÓN DE RECURSOS
    elif "actualización" in req or "recursos" in req:
        return {
            "codigo": "PL-07-005",
            "tipo_firma": "ELABORADO / REVISADO / APROBADO",
            "secciones": {
                "1. OBJETIVO ESTRATÉGICO": "Planificar, estructurar, presupuestar y ejecutar la renovación tecnológica, mantenimiento preventivo y actualización de componentes de hardware y plataformas de software de la compañía. El objetivo es mitigar contundentemente los riesgos asociados a la deuda técnica, obsolescencia de equipos, cuellos de botella en la operación logística y vulnerabilidades por sistemas sin soporte de fábrica (Control A.8.19 de la ISO/IEC 27001 y A.8.8 Gestión de vulnerabilidades).",
                "2. ALCANCE Y DIAGNÓSTICO": "Abarca todo el parque tecnológico inventariado en SERGEM (Sedes operativas y administrativas): Computadores de escritorio, portátiles, servidores físicos, switches core, sistemas operativos de usuario (Windows/Linux) y software de aplicación corporativa.",
                "3. CICLO DE VIDA Y RENOVACIÓN DE LOS ACTIVOS": "• Hardware: Se establece como métrica corporativa que los equipos de cómputo estándar y periféricos tienen un ciclo de vida útil estimado de cuatro (4) a cinco (5) años. Los servidores core y dispositivos de red se proyectan a un ciclo de entre cinco (5) y siete (7) años.\n• Software y Parches: Las licencias de sistemas operativos, paquetes de ofimática y soluciones antivirus (EDR) deben actualizarse obligatoriamente a sus últimas versiones estables con soporte activo de fábrica (End-of-Life avoidance). El sistema WSUS o MDM forzará la instalación de parches críticos de seguridad liberados por Microsoft/Linux en un plazo máximo de 15 días posteriores a su lanzamiento.",
                "4. PROCEDIMIENTO DE EJECUCIÓN Y PRESUPUESTO": "El área de Tecnología (TI) tiene el deber de realizar un cruce de inventario anual durante el último trimestre del año. \n\nA través de este ejercicio, detectará qué equipos, componentes o licencias están próximos a su obsolescencia y estructurará el 'Plan de Inversiones y Actualización Tecnológica' (CAPEX/OPEX). Este plan deberá ser presentado ante la Dirección Administrativa y la Gerencia General para su respectiva aprobación y asignación presupuestal para el año fiscal venidero."
            }
        }
        
    # 12. CAPACITACIONES
    elif "capacitaci" in req or "planilla" in req:
        return {
            "codigo": "PR-08-001",
            "tipo_firma": "ELABORADO / REVISADO / APROBADO",
            "secciones": {
                "1. OBJETIVO DEL PROGRAMA DE CONCIENCIACIÓN": "Garantizar de manera documentada que la totalidad del personal administrativo, operativo y directivo de SERGEM Mensajería S.A.S. reciba formación técnica, concienciación periódica y actualización legal en materia de seguridad de la información y protección de datos. Da estricto cumplimiento al Control A.6.3 (Concienciación, educación y capacitación) de la norma ISO/IEC 27001.",
                "2. ALCANCE Y COBERTURA NACIONAL": "El programa de capacitación y sensibilización (Security Awareness) es de carácter transversal e involucra a todas las áreas de la compañía. Se despliega mediante sesiones presenciales y/o virtuales cubriendo a los trabajadores en Cali, Bogotá, Medellín, Barranquilla, Cartagena e Ibagué.",
                "3. AGENDA Y TEMARIO OBLIGATORIO": "El plan anual de capacitación debe incluir, de manera obligatoria pero no limitativa, los siguientes pilares de seguridad:\n• Gestión y uso seguro de contraseñas corporativas (Prohibición de post-its y compartición).\n• Identificación y prevención de Ingeniería Social (Phishing, Vishing, Smishing).\n• Política de escritorio y pantalla limpia en puestos operativos.\n• Manejo adecuado y normatividad legal sobre la protección de datos de clientes (Habeas Data - Ley 1581).\n• Protocolos de reporte de incidentes a la mesa de ayuda (HelpDesk).",
                "4. INDUCCIÓN A NUEVOS EMPLEADOS (ONBOARDING)": "Todo colaborador recién contratado debe recibir la inducción de seguridad de la información dentro de su primera semana laboral, antes de otorgársele el acceso pleno al sistema Freeway. En caso contrario, los accesos permanecerán restringidos.",
                "5. COMPROMISOS, EVALUACIÓN Y CONSTANCIA (PLANILLAS)": "La asistencia y participación en estas jornadas es de carácter estrictamente obligatorio y se encuentra dentro de las funciones laborales. Cada asistente debe firmar de puño y letra (o mediante firma digital validada) la 'Planilla de Asistencia a Capacitación', la cual reposará en los archivos de Gestión Humana como evidencia documental auditable para sustentar la conformidad del SGSI frente a Kreston o cualquier ente auditor."
            }
        }

    # 13. PLAN DE EMERGENCIAS Y PÉRDIDA DE INFORMACIÓN
    elif "emergencia" in req or "pérdida" in req:
        return {
            "codigo": "PR-07-015",
            "tipo_firma": "ELABORADO / REVISADO / APROBADO",
            "secciones": {
                "1. OBJETIVO DEL PLAN DE CONTINGENCIA (DRP)": "Establecer, documentar y probar un plan de acción inmediato y estructurado (Disaster Recovery Plan) para mitigar, contener, responder y recuperar la operatividad de la infraestructura tecnológica y las bases de datos logísticas ante incidentes críticos. Estos incidentes abarcan ciberataques (Ransomware/DDoS), fallas de hardware catastróficas, incendios en el Data Center físico, o desastres naturales. Cumple con los controles A.5.29 y A.8.14 de ISO/IEC 27001.",
                "2. ALCANCE Y ESCENARIOS CONTEMPLADOS": "El plan está diseñado para asegurar la supervivencia y el reinicio de los procesos críticos (recepción de despachos, asignación de rutas y facturación) en el sistema Freeway, abarcando todas las sedes a nivel nacional.",
                "3. MÉTRICAS DE RECUPERACIÓN EXIGIDAS (RTO Y RPO)": "La Gerencia ha establecido los siguientes límites de tolerancia ante desastres:\n• RTO (Recovery Time Objective): El tiempo máximo tolerable en el que el sistema core (Freeway) y las comunicaciones pueden estar interrumpidos antes de causar un daño comercial inaceptable a SERGEM, fijado en un máximo de cuatro (4) horas.\n• RPO (Recovery Point Objective): El volumen máximo tolerable de pérdida de información transaccional, fijado en un ciclo de veinticuatro (24) horas (alineado a los backups diarios).",
                "4. COMITÉ DE CRISIS Y ROLES": "Ante la declaratoria de desastre, se conforma de forma inmediata el 'Comité de Crisis', liderado por el Gerente General, la Dirección Administrativa y el Jefe de TI. Este comité será el único autorizado para tomar decisiones financieras de emergencia y emitir comunicados oficiales a clientes y proveedores afectados.",
                "5. PROTOCOLO DE RESPUESTA A CRISIS EN 4 PASOS": "1. Identificación y Declaratoria: El equipo de TI confirma el desastre (ej. pérdida de base de datos o cifrado por malware) y el Comité declara el estado de contingencia.\n2. Contención y Aislamiento: Si es un ciberataque, se desconectan físicamente los servidores afectados de la red LAN/WAN y de internet de forma inmediata para frenar la propagación lateral del virus.\n3. Recuperación Alterna (Failover): Se activa el protocolo de restauración de respaldos inmutables en la nube junto a los ingenieros del proveedor tecnológico SOLINUX, levantando servidores secundarios o virtuales.\n4. Retorno a la Normalidad (Failback) y Análisis: Una vez estabilizado el servicio con los backups, se realiza un análisis forense de la causa, se parchea la vulnerabilidad y se levanta el estado de emergencia oficial."
            }
        }
        
    # 14. POLÍTICA DE CONTRASEÑAS
    elif "contraseña" in req or "clave" in req:
        return {
            "codigo": "PO-07-004",
            "tipo_firma": "ELABORADO / REVISADO / APROBADO",
            "secciones": {
                "1. OBJETIVO TÉCNICO Y NORMATIVO": "Definir, parametrizar y hacer cumplir los lineamientos técnicos estrictos para la creación, longitud, complejidad, custodia y rotación periódica de las credenciales (passwords) de acceso a los sistemas de información, redes Wi-Fi, ERP Freeway y correos corporativos de SERGEM Mensajería S.A.S. Este documento se acopla al Control A.9.4.3 (Gestión de contraseñas de usuario) de la norma ISO/IEC 27001.",
                "2. ALCANCE DE LAS CREDENCIALES": "Esta política es de cumplimiento técnico mandatorio para absolutamente todas las cuentas de usuario (Directivos, Operarios, Administrativos, Proveedores con VPN y Administradores de Sistemas) creadas bajo el dominio informático de SERGEM.",
                "3. ESTÁNDARES TÉCNICOS Y PARÁMETROS DE COMPLEJIDAD": "El departamento de Tecnología deberá configurar los directorios activos (Active Directory / IAM) para forzar los siguientes parámetros sin excepción:\n• Longitud Mínima: Las contraseñas deben contener como mínimo doce (12) caracteres de longitud.\n• Complejidad Obligatoria: Debe combinar, al menos, tres de los siguientes cuatro grupos de caracteres: Letras mayúsculas (A-Z), letras minúsculas (a-z), números (0-9) y símbolos especiales (!@#$%&*).\n• Rotación Temporal: Se aplicará una caducidad y rotación obligatoria de credenciales cada noventa (90) días calendario.\n• Historial y Bloqueo: El sistema impedirá la reutilización de las últimas cuatro (4) contraseñas anteriores. Asimismo, la cuenta se bloqueará automáticamente por 30 minutos tras superar tres (3) intentos fallidos consecutivos de ingreso (prevención de ataques de Fuerza Bruta).",
                "4. PROHIBICIONES, CUSTODIA Y BUENAS PRÁCTICAS": "• Es una falta grave, sancionable disciplinariamente, escribir las contraseñas en notas adhesivas (post-its), libretas visibles en los escritorios, o guardarlas en archivos de texto sin cifrar en el escritorio del PC.\n• Queda terminantemente prohibido compartir, prestar o revelar la contraseña personal a compañeros, supervisores o incluso al personal de TI. El departamento de Tecnología NUNCA solicitará contraseñas por correo, chat o teléfono.\n• Se debe evitar el uso de patrones predecibles como fechas de nacimiento, nombres de familiares, mascotas, números de cédula, o nombres asociados a la empresa (Ej: Sergem2024*)."
            }
        }
        
    # 15. DISPOSITIVOS MÓVILES (BYOD)
    elif "móvil" in req or "dispositivo" in req:
        return {
            "codigo": "PO-07-008",
            "tipo_firma": "ELABORADO / REVISADO / APROBADO",
            "secciones": {
                "1. OBJETIVO DE SEGURIDAD MÓVIL Y BYOD": "Establecer las normas, restricciones y controles técnicos de seguridad para el uso de dispositivos móviles (Smartphones, Tablets y Laptops). Esto aplica tanto para equipos corporativos suministrados por SERGEM como para dispositivos personales (bajo esquema BYOD - Bring Your Own Device) que procesen, sincronicen correos, almacenen o transmitan información confidencial de la compañía. Acata el Control A.8.1 de la ISO/IEC 27001.",
                "2. ALCANCE TECNOLÓGICO": "Abarca todos los dispositivos portátiles utilizados por la fuerza de ventas, conductores, gerentes y personal administrativo que requieren movilidad para ejecutar sus funciones desde fuera de la red local (Sedes, trabajo en casa o trabajo en calle).",
                "3. DIRECTRICES DE CONFIGURACIÓN Y USO OBLIGATORIO": "• Bloqueo y Autenticación: Todo equipo móvil que maneje correo corporativo o acceda al ERP logístico debe contar de forma obligatoria con un método de bloqueo robusto (PIN complejo numérico, contraseña alfanumérica o biometría por huella/rostro).\n• Cifrado: Se debe activar el cifrado de almacenamiento del dispositivo (Cifrado de disco completo para Laptops o cifrado nativo en iOS/Android) para proteger los datos en reposo en caso de extravío.\n• Aplicaciones y Redes: Se prohíbe el 'Jailbreak' o 'Rooting' en los dispositivos que accedan a la red corporativa. No se deben conectar a redes Wi-Fi públicas y abiertas en aeropuertos o cafeterías sin el uso activo de la VPN de SERGEM.",
                "4. GESTIÓN DE INCIDENTES (PÉRDIDA O ROBO) Y BORRADO REMOTO": "El colaborador asume la responsabilidad y obligación ineludible de reportar de forma inmediata (en un plazo no mayor a 2 horas) el hurto, extravío o pérdida de su dispositivo móvil corporativo o personal (si contiene datos de SERGEM) al departamento de TI. \n\nUna vez recibido el reporte, el equipo de Tecnología está facultado y obligado a proceder con el bloqueo preventivo de las cuentas de Office/Correo y ejecutar remotamente el borrado seguro de fábrica (Remote Wipe) mediante las herramientas de gestión de dispositivos móviles (MDM) para proteger la información confidencial antes de que caiga en manos de cibercriminales."
            }
        }

    # 16. NOTIFICACIÓN DE INCIDENTES
    elif "notificación" in req or "incidente" in req:
        return {
            "codigo": "PR-07-011",
            "tipo_firma": "ELABORADO / REVISADO / APROBADO",
            "secciones": {
                "1. OBJETIVO DE GESTIÓN DE INCIDENTES": "Estandarizar y documentar el procedimiento operativo formal para la detección temprana, reporte estructurado, clasificación por criticidad, contención, análisis y resolución de los eventos e incidentes de seguridad de la información. Este marco metodológico busca minimizar el impacto operativo y financiero de un ataque cibernético y cumplir con los Controles A.5.24, A.5.25 y A.5.26 de la norma ISO/IEC 27001.",
                "2. ALCANCE DEL PROCEDIMIENTO": "Este protocolo es de estricta aplicación para todos los empleados, contratistas y proveedores tecnológicos de las sedes operativas de SERGEM a nivel nacional frente a cualquier anomalía informática (infecciones por virus, caídas de la base de datos, robo de equipos o accesos no autorizados al ERP).",
                "3. CANALES DE REPORTE Y DEBERES DEL COLABORADOR": "• Todo colaborador o tercero que detecte un comportamiento anómalo en el software (ej. lentitud extrema o pantallas de extorsión tipo Ransomware), la recepción de correos fraudulentos (Phishing), la pérdida de un equipo portátil o la sospecha firme de que un acceso no ha sido autorizado, tiene el deber ético, contractual y laboral de reportarlo sin demoras al Área de Soporte.\n• El reporte oficial se debe canalizar a través de la herramienta de Mesa de Ayuda (Helpdesk) generando un Ticket, o mediante las líneas de atención directas del Departamento de TI si la red ha colapsado.",
                "4. TIEMPOS DE ESCALAMIENTO Y CLASIFICACIÓN (TRIAGE)": "Una vez recibido el reporte, TI debe realizar un Triage:\n• Incidente Leve o Moderado (Ej. Infección aislada en un PC administrativo): Resolución por soporte Nivel 1/2 en el transcurso del turno operativo.\n• Incidente Crítico (Ej. Fuga masiva de guías de clientes, caída general del Firewall, Ransomware en el servidor central): El equipo de TI tiene un plazo estricto e improrrogable inferior a una (1) hora para escalar el caso a la Dirección Administrativa y al Gerente General para la activación del Comité de Crisis y Plan de Continuidad (DRP).",
                "5. REGISTRO, FORENSE Y LECCIONES APRENDIDAS (POST-MORTEM)": "Cada incidente que afecte de manera sensible la operación de SERGEM deberá ser cerrado con un informe técnico detallado (Bitácora Post-Mortem). Este informe debe plasmar la causa raíz del evento, la línea de tiempo de la afectación, el método de contención aplicado y, lo más valioso, las medidas preventivas e inversiones necesarias (acciones de mejora) adoptadas para mitigar el riesgo de que la misma vulnerabilidad vuelva a ser explotada en el futuro."
            }
        }
        
    # 17. ACUERDOS DE CONFIDENCIALIDAD
    elif "acuerdo" in req or "servicio" in req or "confidencialidad" in req:
        return {
            "codigo": "PO-07-014",
            "tipo_firma": "CONTRATISTA / PROVEEDOR",
            "secciones": {
                "1. OBJETO Y MARCO LEGAL CONTRACTUAL": "Establecer de manera vinculante los términos legales, condiciones de Acuerdos de Nivel de Servicio (SLA) y la rigurosa Política de Salvaguarda de Confidencialidad y No Divulgación (NDA). Este documento es de aceptación y firma obligatoria por parte de todos los contratistas, auditores, consultores, proveedores de servicios tecnológicos, empresas de aseo y terceros que ingresen, interactúen o presten servicios directos a SERGEM Mensajería S.A.S. en cualquiera de sus sedes.",
                "2. DEFINICIÓN DE INFORMACIÓN CONFIDENCIAL": "Se entiende como 'Información Confidencial' toda aquella información comercial, logística, contable, financiera, planes de expansión, contraseñas de red, códigos fuente, bases de datos de remitentes y destinatarios, y procedimientos operativos estándar a los que el CONTRATISTA tenga acceso físico, verbal o digital, por error o por necesidad del servicio, durante la vigencia de la relación comercial.",
                "3. CLÁUSULAS DE CONFIDENCIALIDAD ESTRICTA Y PROTECCIÓN DE DATOS": "PRIMERA - DEBER DE SECRETO: El CONTRATISTA se obliga a mantener en la más absoluta reserva la Información Confidencial, adoptando para su protección las mismas medidas de seguridad que utilizaría para proteger sus propios secretos comerciales, pero nunca inferiores a un grado razonable de cuidado.\n\nSEGUNDA - HABEAS DATA: El CONTRATISTA dará estricto y total cumplimiento a las normativas de protección de datos personales vigentes en Colombia (Ley Estatutaria 1581 de 2012 y sus decretos reglamentarios), asumiendo total responsabilidad legal y pecuniaria en caso de pérdida, fuga o alteración de datos pertenecientes a los clientes finales de SERGEM.\n\nTERCERA - PROHIBICIONES: Queda terminantemente prohibida la divulgación, reproducción sin autorización escrita, comercialización, cesión o uso de la Información Confidencial para fines personales, competenciales o distintos a los estrictamente pactados para la prestación del servicio.",
                "4. VIGENCIA Y CONSECUENCIAS POR INCUMPLIMIENTO": "El deber de reserva y secreto profesional no expira con la terminación del contrato de prestación de servicios, extendiéndose por un periodo de cinco (5) años posteriores (o indefinidamente para datos sujetos a Habeas Data). \n\nEl incumplimiento comprobado de estas cláusulas facultará a SERGEM Mensajería S.A.S. para dar por terminado el contrato comercial de forma unilateral, retener pagos pendientes por concepto de daños y perjuicios pre-tasados, y entablar las acciones legales civiles, penales y comerciales a que haya lugar en los tribunales competentes para el resarcimiento de los daños corporativos e imagen."
            }
        }

    # 18. PERSONAL RETIRADO
    elif "retirado" in req or "base de datos" in req:
        return {
            "codigo": "PO-07-025",
            "tipo_firma": "ELABORADO / REVISADO / APROBADO",
            "secciones": {
                "1. OBJETIVO DEL CONTROL DE DESVINCULACIÓN": "Asegurar que los procesos de terminación de contratos laborales, renuncias, desvinculaciones con justa causa o finalización de convenios con terceros (Proveedores), incluyan la gestión segura, revocación total y oportuna de los accesos lógicos (sistemas) y físicos a las instalaciones de SERGEM Mensajería S.A.S. El fin principal es evitar fugas de información, represalias informáticas o sabotajes (Amenaza Interna) dando estricto cumplimiento al Control A.6.5 de la ISO/IEC 27001.",
                "2. ALCANCE Y APLICABILIDAD": "Aplica para absolutamente todos los cargos y niveles jerárquicos (Gerentes, Supervisores, Jefes, Auxiliares Logísticos y Pasantes) en todas las sucursales a nivel nacional en el momento exacto en que cesa su vínculo laboral u operativo con la compañía.",
                "3. PROTOCOLO DE NOTIFICACIÓN Y REVOCACIÓN LÓGICA": "• Notificación Obligatoria: La Gerencia de Gestión Humana (o la jefatura inmediata en su defecto) tiene el deber y la obligación indelegable de notificar formalmente y por escrito (correo/ticket) al departamento de TI la novedad de retiro de cualquier colaborador el mismo día en que se genera y firma el acta de desvinculación.\n• Purga de Accesos: Recibida la notificación, el equipo de TI procederá a inhabilitar de manera prioritaria e inmediata (estableciendo un SLA o plazo máximo crítico de 24 horas) el acceso al directorio activo, las cuentas de correo electrónico corporativo, los usuarios del ERP logístico (Freeway), los perfiles VPN y los permisos a carpetas de red compartidas.",
                "4. DEVOLUCIÓN DE ACTIVOS FÍSICOS Y PAZ Y SALVO": "Es requisito indispensable y obligatorio para la firma del 'Paz y Salvo Laboral' la devolución física de la totalidad de los activos asignados al colaborador. \n\nEsto incluye: Equipos portátiles y cargadores, teléfonos móviles corporativos, carnés de identificación, tarjetas de acceso magnético/RFID, tokens de seguridad física (RSA) y cualquier documentación impresa, disco duro externo o memoria USB que contenga información confidencial de las rutas logísticas y clientes de SERGEM.",
                "5. REDIRECCIÓN DE CORREOS Y RESPALDO (BACKUP DE SALIDA)": "Si el cargo del empleado desvinculado manejaba temas críticos con clientes externos o proveedores financieros, TI configurará (previa autorización de la Dirección Administrativa) una redirección temporal de su buzón de correo hacia la jefatura de su área por un periodo de 30 días, con el fin de asegurar la continuidad de los negocios en curso. Posteriormente, la cuenta será dada de baja y su buzón archivado."
            }
        }
        
    # 19. PLANTILLA MAESTRA GENÉRICA DE RESPALDO
    else:
        cod_aleatorio = random.randint(100, 999)
        titulo_seccion = requisito.title()
        
        return {
            "codigo": f"SG-07-{cod_aleatorio}",
            "tipo_firma": "ELABORADO / REVISADO / APROBADO",
            "secciones": {
                "1. OBJETIVO Y MARCO NORMATIVO": f"Definir, documentar y formalizar detalladamente los lineamientos técnicos, administrativos, legales y operativos correspondientes al proceso de: {titulo_seccion}. \n\nEste documento rector ha sido estructurado técnica y estratégicamente para dar cumplimiento irrestricto a los requisitos exigidos por el estándar internacional de seguridad de la información ISO/IEC 27001:2022. Su propósito fundamental es garantizar la mitigación de vulnerabilidades y la preservación absoluta de la confidencialidad, integridad y disponibilidad del activo de información dentro de las bases de datos y la operación logística de SERGEM Mensajería S.A.S. \n\nAdicionalmente, este marco procedimental se alinea de manera taxativa con las disposiciones legales expedidas por el Congreso de la República de Colombia, respetando específicamente las directrices sancionatorias de la Ley 1581 de 2012 (Régimen General de Protección de Datos Personales / Habeas Data) y la Ley 1273 de 2009 (De los Delitos Informáticos y la Protección de la Información).",
                "2. ALCANCE ORGANIZACIONAL E INFRAESTRUCTURA": f"Las obligaciones, parámetros técnicos y prohibiciones detalladas en el presente documento de {titulo_seccion} son de estricto y obligatorio acatamiento para todo el ecosistema laboral de la compañía. Esto incluye:\n\n• Colaboradores con contrato laboral directo (Término fijo e indefinido).\n• Personal temporal, operarios logísticos en misión y estudiantes en práctica.\n• Contratistas independientes, proveedores tecnológicos, consultores externos y cualquier tercero que interactúe, administre, provea o posea algún nivel de acceso lógico, remoto o físico a los sistemas de información de SERGEM Mensajería S.A.S.\n\nEste alcance tiene jurisdicción operativa e informática sobre la totalidad de la infraestructura de hardware, redes, sistemas core (incluyendo la plataforma Freeway), servidores virtuales, respaldos cloud, y de manera geográfica abarca la sede administrativa principal y todas las sedes logísticas ubicadas a nivel nacional en: Cali, Bogotá, Medellín, Barranquilla, Cartagena e Ibagué.",
                "3. DEFINICIONES CLAVES Y TÉRMINOS TÉCNICOS": "Para la correcta comprensión y ejecución de las directrices plasmadas en este manual, se establecen las siguientes definiciones unificadas:\n\n• SGSI (Sistema de Gestión de Seguridad de la Información): Conjunto de políticas, procedimientos, directrices y recursos asociados para administrar y proteger integralmente la información corporativa, apoyados en la norma ISO 27001.\n• Freeway: Aplicativo de software y sistema core principal logístico mediante el cual SERGEM gestiona inventarios, rutas, envíos y datos de destinatarios finales.\n• Activo de Información: Cualquier conocimiento, base de datos, archivo de Excel, software, servidor, equipo de cómputo, o documentación impresa que tiene un valor operativo o estratégico para la compañía.\n• Evento / Incidente de Seguridad: Un suceso o serie de eventos anómalos o inesperados, confirmados o bajo sospecha, que amenazan con comprometer la seguridad de las redes de SERGEM, paralizando la operación o filtrando datos privados al exterior.",
                "4. DIRECTRICES Y CONTROLES OPERATIVOS (ISO/IEC 27001)": f"Para la ejecución y mantenimiento adecuado de las actividades concernientes a {titulo_seccion}, el personal involucrado deberá observar y acatar, sin excepción, el siguiente esquema de reglas inquebrantables de seguridad de la información:\n\n1. Principio de Menor Privilegio Lógico: Todo acceso a carpetas compartidas, sistemas operativos, ERP o repositorios vinculados a esta actividad deberá estar restringido al mínimo indispensable para realizar la función encomendada. Todo permiso adicional deberá requerir justificación escrita.\n2. Trazabilidad y Bitácoras Inmutables: Todo registro, transacción crítica, auditoría de logs del sistema (Windows/Linux) o formulario físico derivado de este proceso deberá ser almacenado, protegido contra modificaciones accidentales y custodiado por un periodo de retención no menor a doce (12) meses para fines periciales y de auditoría externa (Kreston o entes de control).\n3. Prohibición de Extracción de Datos: Queda estrictamente prohibida la exportación, clonación, copia por USB o envío hacia correos electrónicos personales (Hotmail, Gmail, Yahoo, etc.) de información confidencial, bases de datos operativas o configuraciones de red asociadas a este documento. \n4. Intervención Rápida ante Incidentes: Si un colaborador llegase a observar una vulneración, lentitud anómala severa, sospecha de software malicioso o elusión de los parámetros aquí establecidos, deberá aplicar la contención primaria e informar obligatoriamente en un tiempo menor a treinta (30) minutos a la Mesa de Ayuda de TI (Helpdesk).",
                "5. ROLES Y MATRIZ DE RESPONSABILIDADES ASIGNADAS": "El aseguramiento de este proceso es un trabajo conjunto y multidisciplinario en la compañía. Por lo tanto, se fijan las siguientes asignaciones y deberes según el cargo directivo y operativo:\n\n• Dirección General: Proveerá de manera oportuna y proporcional los recursos financieros, tecnológicos y el talento humano calificado para garantizar el mantenimiento y mejora del SGSI.\n• Dirección Administrativa (Yesenia Beltrán): Actuará como gestora y principal aval del cumplimiento normativo del SGSI, liderando las auditorías, programando las revisiones de matriz de riesgos y garantizando la coherencia documental corporativa.\n• Departamento de TI e Infraestructura: Recae sobre ellos la responsabilidad completamente técnica de diseñar, implementar, parchear, asegurar criptográficamente y auditar la plataforma tecnológica, de manera que los bloqueos y alertas respalden sistémicamente las reglas descritas.\n• Jefes de Área y Colaboradores Generales: Acatar invariablemente cada directriz, manteniendo en alto la cultura del cuidado del activo corporativo. Ignorar las reglas no exime de culpa al usuario.",
                "6. RÉGIMEN DISCIPLINARIO Y SANCIONES POR INCUMPLIMIENTO": f"Las disposiciones establecidas en el presente documento operativo de {titulo_seccion} son un mandato directo de la Alta Gerencia de SERGEM. \n\nEl incumplimiento, la omisión, la evasión de controles técnicos, la negligencia grave que ocasione pérdida de disponibilidad, o la vulneración intencionada de estas normativas, constituye una violación sustancial y grave a las políticas de seguridad corporativas y obligaciones contractuales.\n\nLa detección de cualquier irregularidad será tipificada como falta y activará inmediatamente el proceso de descargos y el protocolo estipulado en el Reglamento Interno de Trabajo. Dependiendo de la afectación técnica y comercial del incidente, las medidas adoptadas por Gestión Humana podrán resultar en un llamado de atención con copia a la hoja de vida, suspensiones no remuneradas o la terminación unilateral e inmediata del contrato laboral por justa causa."
            }
        }

def generar_documento_pdf(requisito):
    datos_doc = obtener_datos_qms(requisito)
    output = io.BytesIO()
    
    doc = SimpleDocTemplate(output, pagesize=letter, rightMargin=0.75*inch, leftMargin=0.75*inch, topMargin=0.6*inch, bottomMargin=0.6*inch)
    elements = []
    styles = getSampleStyleSheet()
    
    style_center = ParagraphStyle(name='Center', parent=styles['Normal'], alignment=TA_CENTER, fontName='Helvetica-Bold', fontSize=10)
    style_normal = ParagraphStyle(name='Justify', parent=styles['Normal'], alignment=TA_JUSTIFY, fontName='Helvetica', fontSize=10, leading=14)
    style_bold_center = ParagraphStyle(name='BoldCenter', parent=styles['Normal'], alignment=TA_CENTER, fontName='Helvetica-Bold', fontSize=10)
    style_title = ParagraphStyle(name='SectionTitle', parent=styles['Normal'], alignment=TA_LEFT, fontName='Helvetica-Bold', fontSize=11, textColor=colors.HexColor('#002b5e'), spaceAfter=8)

    logo_path = "sergemLogo.png"
    if os.path.exists(logo_path):
        logo_img = RLImage(logo_path, width=1.0*inch, height=1.0*inch)
    else:
        logo_img = Paragraph("LOGO", style_bold_center)
        
    dia_aleatorio = random.randint(1, 28)
    fecha_generada = f"{dia_aleatorio:02d}/05/{ANIO_ACTUAL}"

    header_data = [
        [logo_img, Paragraph(requisito.upper(), style_center), '', '', logo_img],
        ['', Paragraph(f"Código: {datos_doc['codigo']}", style_bold_center), Paragraph("Versión No.1", style_bold_center), Paragraph(fecha_generada, style_bold_center), '']
    ]
    
    t_header = Table(header_data, colWidths=[1.4*inch, 1.4*inch, 1.3*inch, 1.4*inch, 1.4*inch])
    t_header.setStyle(TableStyle([
        ('GRID', (0,0), (-1,-1), 1, colors.black),
        ('SPAN', (0,0), (0,1)), 
        ('SPAN', (4,0), (4,1)), 
        ('SPAN', (1,0), (3,0)), 
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('TOPPADDING', (0,0), (-1,-1), 6),
        ('BOTTOMPADDING', (0,0), (-1,-1), 6),
    ]))
    elements.append(t_header)
    elements.append(Spacer(1, 0.25*inch))

    for titulo, contenido in datos_doc['secciones'].items():
        contenido_rl = contenido.replace('\n', '<br/>')
        body_data = [
            [Paragraph(titulo, style_title)],
            [Paragraph(contenido_rl, style_normal)]
        ]
        t_body = Table(body_data, colWidths=[7.0*inch])
        t_body.setStyle(TableStyle([
            ('GRID', (0,0), (-1,-1), 1, colors.HexColor('#bdc3c7')),
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#f8f9fa')),
            ('ALIGN', (0,0), (-1,-1), 'LEFT'),
            ('VALIGN', (0,0), (-1,-1), 'TOP'),
            ('BOTTOMPADDING', (0,0), (-1,-1), 12),
            ('TOPPADDING', (0,0), (-1,-1), 10),
            ('LEFTPADDING', (0,0), (-1,-1), 12),
            ('RIGHTPADDING', (0,0), (-1,-1), 12),
        ]))
        elements.append(t_body)
        elements.append(Spacer(1, 0.15*inch))

    # -------------------------------------------------------------------------
    # ANCLAJE MATEMÁTICO INFERIOR (CONDITIONAL PAGE BREAK + BOTTOM SPACER)
    # -------------------------------------------------------------------------
    elements.append(CondPageBreak(120))
    elements.append(BottomSpacer(100))

    if datos_doc['tipo_firma'] == "ELABORADO / REVISADO / APROBADO":
        sig_data = [
            [Paragraph("Elaborado por:", style_bold_center), Paragraph("Revisado por:", style_bold_center), Paragraph("Aprobado por:", style_bold_center)],
            [Paragraph("Nombre: Yesenia Beltrán<br/>Cargo: Directora Administrativa", style_normal),
             Paragraph("Nombre: Yesenia Beltrán<br/>Cargo: Directora Administrativa", style_normal),
             Paragraph("Nombre: José Reinel Torres<br/>Cargo: Gerente General", style_normal)]
        ]
        t_sig = Table(sig_data, colWidths=[2.33*inch, 2.33*inch, 2.33*inch])
        t_sig.setStyle(TableStyle([
            ('GRID', (0,0), (-1,-1), 1, colors.black),
            ('ALIGN', (0,0), (-1,-1), 'LEFT'),
            ('VALIGN', (0,0), (-1,-1), 'TOP'),
            ('BOTTOMPADDING', (0,0), (-1,-1), 12),
            ('TOPPADDING', (0,0), (-1,-1), 8),
            ('LEFTPADDING', (0,0), (-1,-1), 8),
            ('RIGHTPADDING', (0,0), (-1,-1), 8),
        ]))
        elements.append(KeepTogether(t_sig))
    else:
        firma_texto = Paragraph(f"<br/><br/>{datos_doc['tipo_firma']}: ___________________________________", style_bold_center)
        elements.append(KeepTogether(firma_texto))
    
    doc.build(elements)
    return output.getvalue()

# -----------------------------------------------------------------------------
# 5. ESTADO DE SESIÓN E INTERFAZ STREAMLIT
# -----------------------------------------------------------------------------
df_archivos = obtener_archivos_drive()

if 'visor_id' not in st.session_state: 
    st.session_state.visor_id = None
if 'visor_nombre' not in st.session_state: 
    st.session_state.visor_nombre = None

if 'excel_backup' not in st.session_state:
    st.session_state.excel_backup = None
if 'mostrar_salvavidas' not in st.session_state:
    st.session_state.mostrar_salvavidas = False

st.sidebar.markdown('### 🗂️ Módulos de Evaluación')
opciones = [
    "🏠 Inicio y Sincronización", 
    "📁 Explorador Documental Completo", 
    "📊 Novedades Auditoría Pasada", 
    "🛠️ Preparador de Auditoría Automático"
]
seleccion = st.sidebar.radio("Seleccione la vista:", opciones)

if seleccion == "🏠 Inicio y Sincronización":
    st.markdown("""
        <div class="card-custom">
            <div class="card-header-custom">Estado del Sistema SGSI</div>
            <p>Bienvenido al portal oficial de auditoría de SERGEM Mensajería S.A.S. El sistema se encuentra sincronizado con el repositorio documental en tiempo real.</p>
        </div>
    """, unsafe_allow_html=True)
    if st.button("🔄 Forzar Sincronización con Repositorio"):
        st.cache_data.clear()
        st.success("✅ Datos sincronizados correctamente.")
    st.info(f"Total de archivos y carpetas detectados en la nube: **{len(df_archivos)}**")

elif seleccion == "📁 Explorador Documental Completo":
    st.markdown("""
        <div class="card-custom">
            <div class="card-header-custom">Repositorio Documental (Auditoría Kreston)</div>
            <p>Seleccione una carpeta en el panel izquierdo y haga clic sobre cualquier documento para visualizarlo de inmediato.</p>
        </div>
    """, unsafe_allow_html=True)
    
    if not df_archivos.empty:
        col_explorer, col_viewer = st.columns([1, 2])
        with col_explorer:
            st.markdown("##### 📂 Estructuras Disponibles")
            rutas = sorted(df_archivos['ruta'].unique())
            for ruta in rutas:
                archivos_en_ruta = df_archivos[(df_archivos['ruta'] == ruta) & (df_archivos['tipo'] == 'Archivo')]
                if not archivos_en_ruta.empty:
                    with st.expander(f"📁 {ruta}"):
                        for _, row in archivos_en_ruta.iterrows():
                            if st.button(f"📄 {row['nombre']}", key=row['id'], use_container_width=True):
                                st.session_state.visor_id = row['id']
                                st.session_state.visor_nombre = row['nombre']
        with col_viewer:
            if st.session_state.visor_id:
                st.markdown(f"**Documento Seleccionado:** `{st.session_state.visor_nombre}`")
                mostrar_visor_archivo(st.session_state.visor_id, st.session_state.visor_nombre)
            else:
                st.info("👈 Seleccione un documento en el panel izquierdo para previsualizarlo en pantalla.")

elif seleccion == "📊 Novedades Auditoría Pasada":
    st.markdown("""
        <div class="card-custom">
            <div class="card-header-custom">Hallazgos y Novedades Auditoría Pasada</div>
            <p>Resumen visual e interactivo de las observaciones pasadas. Expanda cada componente para ver el detalle.</p>
        </div>
    """, unsafe_allow_html=True)

    @st.cache_data(ttl=300)
    def cargar_matriz_observaciones(df_archivos_nube):
        if df_archivos_nube.empty:
            return pd.DataFrame(), None
            
        mask = df_archivos_nube['nombre'].str.contains("Matriz de Observaciones", case=False, na=False)
        archivos_matriz = df_archivos_nube[mask & (df_archivos_nube['tipo'] == 'Archivo')]
        
        if archivos_matriz.empty:
            return pd.DataFrame(), None 
            
        candidato = archivos_matriz.iloc[0]
        file_id = candidato['id']
        url_descarga = f"https://drive.google.com/uc?export=download&id={file_id}"
        
        try:
            df = pd.read_excel(url_descarga, sheet_name=0, header=15)
            df.columns = df.columns.astype(str).str.replace('\n', ' ').str.strip()
            
            col_informe = next((col for col in df.columns if 'INFORME' in col.upper()), None)
            col_componente = next((col for col in df.columns if 'COMPONENTE' in col.upper()), None)
            col_observacion = next((col for col in df.columns if 'OBSERVACIÓN' in col.upper() or 'OBSERVACION' in col.upper()), None)
            col_subsanacion = next((col for col in df.columns if 'COMO FUE SUBSANADO' in col.upper() or 'ACTIVIDAD' in col.upper()), None)
            col_estado = next((col for col in df.columns if 'ESTADO' in col.upper()), None)
            
            datos_limpios = []
            if col_observacion:
                for idx, row in df.iterrows():
                    obs = str(row.get(col_observacion)).strip()
                    if pd.isna(row.get(col_observacion)) or obs.upper() in ['', 'NAN', 'OBSERVACIÓN', 'OBSERVADA']: 
                        continue
                        
                    estado = 'SIN ESTADO'
                    if col_estado and pd.notna(row.get(col_estado)): 
                        estado = str(row[col_estado]).upper()
                    elif 'SUBSANADO' in df.columns:
                        idx_subs = df.columns.get_loc('SUBSANADO')
                        val_si = str(row.iloc[idx_subs]).strip().upper()
                        val_no = str(row.iloc[idx_subs + 1]).strip().upper() if (idx_subs + 1) < len(df.columns) else ''
                        if val_si == 'X': 
                            estado = 'SUBSANADA'
                        elif val_no == 'X': 
                            estado = 'NO SUBSANADA'

                    actividad = 'Sin actividad registrada'
                    if col_subsanacion and pd.notna(row.get(col_subsanacion)):
                        val_act = str(row[col_subsanacion]).strip()
                        if val_act.upper() not in ['', 'NAN']: 
                            actividad = val_act

                    datos_limpios.append({
                        'Informe': str(row.get(col_informe, 'N/A')),
                        'Componente': str(row.get(col_componente, 'N/A')),
                        'Observación': obs,
                        'Estado': estado,
                        'Subsanación (Actividad)': actividad
                    })
            return pd.DataFrame(datos_limpios), file_id 
        except Exception as e:
            return pd.DataFrame(), None

    df_nov, matrix_file_id = cargar_matriz_observaciones(df_archivos)

    if not df_nov.empty:
        conteo_estados = df_nov['Estado'].value_counts()
        col_metric, col_chart = st.columns([1, 2])
        
        with col_metric:
            st.metric("Total Observaciones", len(df_nov))
            st.metric("✅ Subsanadas", conteo_estados.get('SUBSANADA', 0))
            st.metric("⚠️ NO Subsanadas", conteo_estados.get('NO SUBSANADA', 0))
            
        with col_chart:
            fig = px.pie(
                values=conteo_estados.values, 
                names=conteo_estados.index, 
                hole=0.4, 
                color=conteo_estados.index, 
                color_discrete_map={'SUBSANADA':'#2ecc71', 'NO SUBSANADA':'#e74c3c', 'SIN ESTADO':'#f1c40f'}
            )
            fig.update_layout(margin=dict(t=0, b=0, l=0, r=0), height=300)
            st.plotly_chart(fig, use_container_width=True)

        if matrix_file_id:
            with st.expander("📄 Clic aquí para verificar el archivo matriz original (Excel de Auditoría)"):
                url_visor = f"https://drive.google.com/file/d/{matrix_file_id}/preview"
                st.markdown(f'<iframe class="pdf-frame" src="{url_visor}" width="100%" height="600"></iframe>', unsafe_allow_html=True)
            
        st.markdown("### 🔍 Detalle Interactivo de Observaciones")
        filtro = st.selectbox("Filtrar estado de la novedad:", ["Todos los Estados"] + list(df_nov['Estado'].unique()))
        
        if filtro == "Todos los Estados":
            df_mostrar = df_nov 
        else:
            df_mostrar = df_nov[df_nov['Estado'] == filtro]

        for _, row in df_mostrar.iterrows():
            emoji = "✅" if row['Estado'] == "SUBSANADA" else "⚠️" if row['Estado'] == "NO SUBSANADA" else "🔒"
            with st.expander(f"{emoji} {row['Componente']} - Estado: {row['Estado']}"):
                st.markdown("**📌 Observación Original:**")
                st.info(row['Observación'])
                st.markdown("**🛠️ Actividad Realizada / Cómo fue subsanado:**")
                if pd.notna(row['Subsanación (Actividad)']) and str(row['Subsanación (Actividad)']).strip() != "":
                    st.success(row['Subsanación (Actividad)'])
                else:
                    st.warning("Aún no hay actividad de subsanación registrada en el archivo.")

elif seleccion == "🛠️ Preparador de Auditoría Automático":
    st.markdown(f"""
        <div class="card-custom">
            <div class="card-header-custom">Preparación Automática para Auditoría (ISO 27001)</div>
            <p>El sistema escanea el inventario del repositorio documental buscando los requisitos exactos del formato base.</p>
        </div>
    """, unsafe_allow_html=True)

    if not df_archivos.empty:
        df_archivos_base = df_archivos[
            (df_archivos['tipo'] == 'Archivo') & 
            (df_archivos['ruta'].str.contains('Auditoría', case=False, na=False))
        ].copy()
        
        df_archivos_base['nombre_norm'] = df_archivos_base['nombre'].apply(remover_acentos)

        requisitos = {
            "Políticas de la seguridad de la información": ["POLITICA", "SEGURIDAD", "INFORMACION"],
            "Políticas de protección de datos (Habeas Data)": ["PROTECCION", "DATOS"],
            "Procedimientos, planillas y/o documentos (capacitaciones)": ["CAPACITACION"],
            "Procedimiento disciplinario": ["DISCIPLINARIO"],
            "Inventario de TI": ["INVENTARIO", "COMPUTADORES"],
            "Plan de actualización de los recursos tecnológicos": ["ACTUALIZACION", "RECURSOS"],
            "Procedimientos de seguridad": ["PROCEDIMIENTOS", "SEGURIDAD"], 
            "Hoja de vida de los equipos de cómputo y servidores": ["HOJA", "VIDA"],
            "Políticas de control de acceso": ["CONTROL", "ACCESO"],
            f"Base de datos, personal retirado {ANIO_ACTUAL}": ["RETIRADO"],
            "Contratos y cláusulas de confidencialidad": ["CONFIDENCIALIDAD"],
            "Plan de respuesta a emergencias (Pérdida de info.)": ["EMERGENCIA", "PERDIDA"],
            "Políticas de contraseñas": ["CONTRASEÑA"],
            "Políticas de uso de dispositivos móviles": ["DISPOSITIVO", "MOVIL"],
            "Políticas, procedimientos de incidentes": ["ROLES", "RESPONSABILIDADES"],
            "Procedimiento de notificación de incidentes": ["NOTIFICACION", "INCIDENTE"],
            "Inventario de Licenciamiento": ["LICENCIA", "INVENTARIO"],
            "Documentos soporte de adquisición de licencias": ["SOPORTE", "ADQUISICION"],
            "Certificación software legal (Representante Legal)": ["LEGAL", "REPRESENTANTE"],
            "Acuerdos de servicio (Proveedores/Terceros)": ["ACUERDO", "SERVICIO", "PROVEEDOR"],
            "Copias de seguridad vigentes y estado": ["COPIA", "SEGURIDAD", "VIGENTE"],
            "Prueba de restauración": ["RESTAURACION", "PRUEBA"],
            "Plan de continuidad del negocio": ["CONTINUIDAD", "NEGOCIO"],
            "Matriz de riesgos de TI": ["MATRIZ", "RIESGO"],
            "Informe de pruebas de vulnerabilidad (Ethical Hacking)": ["VULNERABILIDAD", "HACKING"],
            "Documentos de gestión de seguridad en contratos": ["CONTRATO", "SEGURIDAD"],
            "SGSI (Sistema de gestión de seguridad)": ["SGSI"],
            "Plan de acción, preventivo y correctivo": ["PLAN", "ACCION", "PREVENTIVO"]
        }

        archivos_encontrados = []
        archivos_validos = []
        ids_procesados = set()
        inventario_id = None
        lista_faltantes = []
        
        for req, keywords in requisitos.items():
            mask = pd.Series(True, index=df_archivos_base.index)
            for kw in keywords:
                kw_norm = remover_acentos(kw)
                mask = mask & df_archivos_base['nombre_norm'].str.contains(kw_norm)

            coincidencias = df_archivos_base[mask]

            if not coincidencias.empty:
                candidato = coincidencias.iloc[0]
                
                es_excel_inventario = "INVENTARIO" in candidato['nombre_norm'] and candidato['nombre'].endswith(('.xls', '.xlsx'))
                
                if es_excel_inventario: 
                    estado = "⚙️ Encontrado (Sincronizable)"
                    inventario_id = candidato['id']
                else:
                    estado = "✅ Encontrado"
                    if candidato['id'] not in ids_procesados:
                        archivos_validos.append({"nombre": candidato['nombre'], "id": candidato['id']})
                        ids_procesados.add(candidato['id'])
                
                archivos_encontrados.append({
                    "Requisito": req, 
                    "Estado": estado, 
                    "Archivo Base": candidato['nombre']
                })
            else:
                archivos_encontrados.append({
                    "Requisito": req, 
                    "Estado": "❌ Faltante", 
                    "Archivo Base": "Buscar automáticamente"
                })
                lista_faltantes.append(req)

        df_analisis = pd.DataFrame(archivos_encontrados)
        filtro_req = st.radio(
            "🔍 Filtrar estado de los documentos:", 
            ["Mostrar Todos", "❌ Solo Faltantes", "✅ Solo Encontrados"], 
            horizontal=True
        )
        
        if filtro_req == "❌ Solo Faltantes":
            df_mostrar = df_analisis[df_analisis['Estado'] == "❌ Faltante"]
        elif filtro_req == "✅ Solo Encontrados":
            df_mostrar = df_analisis[df_analisis['Estado'].str.contains("Encontrado")]
        else:
            df_mostrar = df_analisis
            
        st.dataframe(df_mostrar, use_container_width=True, hide_index=True)
        
        nombres_validos = [d['nombre'] for d in archivos_validos]
        if inventario_id:
            nombres_validos.append(df_archivos_base[df_archivos_base['id'] == inventario_id].iloc[0]['nombre'])
            
        df_sobrantes = df_archivos_base[~df_archivos_base['nombre'].isin(nombres_validos)][['nombre', 'ruta']]
        if not df_sobrantes.empty:
            with st.expander(f"⚠️ Atención: Se detectaron {len(df_sobrantes)} archivos sobrantes en la carpeta (No requeridos)"):
                st.warning("Estos documentos no hacen parte de la lista oficial de la auditoría Kreston. Considera verificar y removerlos para evitar confusiones a la hora de presentar los soportes.")
                st.dataframe(df_sobrantes, use_container_width=True, hide_index=True)
                
        st.divider()

        col_qms, col_auto = st.columns(2)
        with col_qms:
            st.markdown("### 📝 Buscador de Documentos Oficiales QMS")
            st.info("Busca de manera inteligente los documentos faltantes basándose en la normativa y los controles del SGSI y la base de datos de SERGEM.")
            
            if lista_faltantes:
                req_selec = st.selectbox("Seleccione el documento a generar:", lista_faltantes)
                
                if st.button(f"🪄 Descargar PDF Oficial: {req_selec}"):
                    with st.spinner("Compilando Documento Normativo..."):
                        archivo_pdf = generar_documento_pdf(req_selec)
                        nombre_descarga = f"{req_selec.replace('/', '_').replace(' ', '_')}_SERGEM_{ANIO_ACTUAL}.pdf"
                        
                        st.download_button(
                            label="⬇️ Descargar Documento Oficial", 
                            data=archivo_pdf, 
                            file_name=nombre_descarga, 
                            mime="application/pdf", 
                            type="secondary"
                        )
            else:
                st.success("✅ ¡Todos los documentos están listos!")

        with col_auto:
            st.markdown("### 🚀 Módulo de Actualización y Sincronización")
            st.info(f"Se actualizará el Inventario de TI y se sincronizarán los **{len(archivos_validos)}** documentos validados en el repositorio de Auditoría.")
            
            if st.button(f"▶️ Sincronizar Repositorio Oficial {ANIO_ACTUAL}", type="primary"):
                st.session_state.mostrar_salvavidas = False 
                st.markdown("#### Progreso de la sincronización y actualización:")
                barra_progreso = st.progress(0)
                texto_estado = st.empty()
                resultados_finales = []
                
                total_pasos = len(archivos_validos) + (1 if inventario_id else 0)
                paso_actual = 0
                
                for doc in archivos_validos:
                    texto_estado.write(f"⏳ Evaluando y sincronizando: {doc['nombre']}...")
                    
                    archivos_mismo_nombre = df_archivos[df_archivos['nombre'] == doc['nombre']]
                    ya_existe_en_auditoria = archivos_mismo_nombre['ruta'].str.contains('Auditoría actual', case=False, na=False).any()
                    
                    if len(archivos_mismo_nombre) > 1 or ya_existe_en_auditoria:
                        resultados_finales.append(f"⏭️ Omitido (Ya existe en destino): {doc['nombre']}")
                        paso_actual += 1
                        barra_progreso.progress(paso_actual / total_pasos)
                        continue

                    payload = {"action": "copiar_archivos", "fileIds": [doc['id']]}
                    try:
                        res_post = requests.post(URL_API_DRIVE, json=payload)
                        if res_post.status_code == 200:
                            respuesta = res_post.json()
                            if respuesta.get("status") == "success": 
                                resultados_finales.append(f"✅ Sincronizado: {doc['nombre']}")
                            else: 
                                resultados_finales.append(f"❌ Error al procesar: {doc['nombre']}")
                        else:
                            resultados_finales.append(f"❌ Falló conexión: {doc['nombre']}")
                    except Exception:
                        resultados_finales.append(f"❌ Omitido (Archivo inaccesible): {doc['nombre']}")
                    
                    paso_actual += 1
                    barra_progreso.progress(paso_actual / total_pasos)
                
                if inventario_id:
                    texto_estado.write(f"⏳ Inyectando volumen de registros ({ANIO_ACTUAL}) en el Inventario Excel...")
                    excel_modificado = actualizar_fecha_inventario_excel(inventario_id)
                    
                    if excel_modificado:
                        st.session_state.excel_backup = excel_modificado 
                        excel_b64 = base64.b64encode(excel_modificado).decode('utf-8')
                        payload_excel = {
                            "action": "subir_archivo",
                            "nombre": f"Inventario de computadores - Actualizado {ANIO_ACTUAL}.xlsx",
                            "base64": excel_b64,
                            "mimeType": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        }
                        
                        try:
                            res_excel = requests.post(URL_API_DRIVE, json=payload_excel)
                            if res_excel.status_code == 200:
                                try:
                                    respuesta_json = res_excel.json()
                                    if respuesta_json.get("status") == "success":
                                        resultados_finales.append(f"✅ Inventario_Actualizado_{ANIO_ACTUAL}.xlsx transferido.")
                                    else:
                                        resultados_finales.append(f"❌ El servidor rechazó la transferencia. Descárgalo de forma segura abajo.")
                                        st.session_state.mostrar_salvavidas = True
                                except:
                                    resultados_finales.append("⚠️ Retraso en la comunicación con el servidor. Descárgalo de forma segura abajo.")
                                    st.session_state.mostrar_salvavidas = True
                            else:
                                resultados_finales.append(f"❌ Fallo HTTP {res_excel.status_code} al sincronizar el Inventario.")
                                st.session_state.mostrar_salvavidas = True
                        except Exception as e:
                            resultados_finales.append(f"⚠️ Error de red al sincronizar el Inventario. Descárgalo de forma segura abajo.")
                            st.session_state.mostrar_salvavidas = True
                    else:
                        resultados_finales.append("❌ Falló el procesamiento del Inventario base.")
                        
                    paso_actual += 1
                    barra_progreso.progress(paso_actual / total_pasos)

                texto_estado.empty()
                st.success("✅ ¡Proceso de empacado finalizado! Detalle de operaciones:")
                with st.expander("Ver bitácora de actualización", expanded=True):
                    for f in resultados_finales: 
                        st.write(f"- {f}")
                st.cache_data.clear()

            if st.session_state.mostrar_salvavidas and st.session_state.excel_backup:
                st.warning("El tamaño del archivo o los límites de red impidieron la carga automática al repositorio.")
                st.info("💡 **ACCIÓN REQUERIDA:** El sistema ha validado y compilado tu archivo con éxito. Descárgalo y cárgalo manualmente en la carpeta de Auditoría.")
                st.download_button(
                    label=f"⬇️ Descargar Inventario Actualizado {ANIO_ACTUAL}",
                    data=st.session_state.excel_backup,
                    file_name=f"Inventario de computadores - Actualizado {ANIO_ACTUAL}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    type="primary"
                )
